"""
Gerador de relatório Excel — CoreExtract.

Abas produzidas:
  1. Dashboard           — KPIs executivos + 4 gráficos semânticos + matriz
  2. Candidatos          — tabela de perfil + fit qualitativo
  3. Pontos_Fortes       — tabela normalizada (1 linha por skill, quando há fit)
  4. Lacunas             — tabela normalizada (1 linha por lacuna, quando há fit)
  5. Mapa de Candidatos  — análise qualitativa, cor por recomendação
  6. Inteligência da Vaga (apenas quando há fit)

Filosofia: cores semânticas obrigatórias (verde=positivo, âmbar=neutro,
vermelho=negativo, azul=identificação). Nenhuma cor arbitrária.
"""
from __future__ import annotations

import io
from collections import Counter
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint as _DP
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Paleta semântica ──────────────────────────────────────────────────────────
_NAVY    = "0D1B2A"
_BLUE    = "1A6BFF"
_CYAN    = "00C8FF"
_GREEN   = "00C07A"      # positivo (Avançar, pontos fortes)
_GREEN_D = "009060"      # borda verde
_AMBER   = "FFB830"      # neutro (Em análise)
_AMBER_D = "CC9020"
_RED     = "FF5757"      # negativo (Não recomendado, lacunas)
_RED_D   = "CC2200"
_PURPLE  = "5A3E8A"      # highlight especial (score médio)
_SLATE   = "F0F4FF"
_WHITE   = "FFFFFF"
_GRAY    = "6B7A99"
_LIGHTBG = "EEF3FF"

_REC_STYLE = {
    "Avançar":         ("D6FAE8", "007A4C"),
    "Em análise":      ("FFF8E1", "7A4F00"),
    "Não recomendado": ("FFE8E8", "CC2200"),
}
_NIVEL_STYLE = {
    "Sênior":       ("D6FAE8", "007A4C"),
    "Especialista": ("E0F8FF", "005580"),
    "Pleno":        ("EEF4FF", "1A3A8B"),
    "Júnior":       ("FFF8E1", "7A4F00"),
}
_REC_ORDER    = ["Avançar", "Em análise", "Não recomendado"]
_NIVEL_ORDER  = ["Especialista", "Sênior", "Pleno", "Júnior", "Indefinido"]
_REC_COLORS   = {
    "Avançar":         _GREEN,
    "Em análise":      _AMBER,
    "Não recomendado": _RED,
}


# ── Helpers de estilo ─────────────────────────────────────────────────────────
def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=11, color=_NAVY, italic=False) -> Font:
    return Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left(wrap=True) -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def _border() -> Border:
    s = Side(style="thin", color="D0D8F0")
    return Border(left=s, right=s, top=s, bottom=s)

def _header_cell(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = _fill(_NAVY);  c.font = _font(bold=True, size=10, color=_WHITE)
    c.alignment = _center();  c.border = _border()
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c

def _kpi_card(ws, row, col1_letter, col2_letter, label, value, bg, fg):
    """Card KPI de 2 colunas × 2 linhas."""
    ws.merge_cells(f"{col1_letter}{row}:{col2_letter}{row}")
    ws.merge_cells(f"{col1_letter}{row+1}:{col2_letter}{row+1}")
    lc = ws[f"{col1_letter}{row}"];   lc.value = label
    vc = ws[f"{col1_letter}{row+1}"]; vc.value = value
    for cell, bold, sz in [(lc, False, 9), (vc, True, 22)]:
        cell.fill = _fill(bg);  cell.font = _font(bold=bold, size=sz, color=fg)
        cell.alignment = _center();  cell.border = _border()


# ── Helpers de dados ──────────────────────────────────────────────────────────
def _fit(r):      return (r.get("resume") or {}).get("fit") or {}
def _resume(r):   return r.get("resume") or {}
def _contato(r):  return (_resume(r).get("contato") or {})
def _habs(r):     return (_resume(r).get("habilidades") or {})
def _exps(r):     return (_resume(r).get("experiencias") or [])

def _score(r) -> int:
    return int((_resume(r).get("scores") or {}).get("score_geral", 0) or 0)

def _cand_name(r, idx: int = 0) -> str:
    return ((_resume(r).get("nome") or r.get("filename", f"C{idx+1}"))[:24])

def _agg(ok_results, fit_field, top=10):
    counts: Counter = Counter()
    for r in ok_results:
        for item in _fit(r).get(fit_field, []):
            item = item.strip()
            if item:
                counts[item] += 1
    return counts.most_common(top)

def _rec_count(ok_results):
    c = {k: 0 for k in _REC_ORDER}
    for r in ok_results:
        rec = _fit(r).get("recomendacao", "")
        if rec in c:
            c[rec] += 1
    return c

def _all_skills(ok_results, top=12):
    counts: Counter = Counter()
    for r in ok_results:
        for s in _habs(r).get("tecnicas", []):
            counts[s.strip()] += 1
    return counts.most_common(top)

def _set_bar_color(series, hex_fill, hex_line=None):
    """Define cor uniforme em uma série de BarChart."""
    series.spPr.solidFill = hex_fill
    series.spPr.ln.solidFill = hex_line or hex_fill

def _color_bar_by_rec(chart_series, sorted_list_with_rec):
    """Colore cada barra individualmente pela recomendação do candidato."""
    for idx, (_, rec) in enumerate(sorted_list_with_rec):
        pt = _DP(idx=idx)
        pt.spPr.solidFill = _REC_COLORS.get(rec, _BLUE)
        chart_series.dPt.append(pt)


# ── Ponto de entrada ──────────────────────────────────────────────────────────
def generate_excel_report(
    results: list[dict],
    tema: str = "Triagem Geral",
    nome_recrutador: Optional[str] = None,
    empresa_recrutadora: Optional[str] = None,
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    ok      = [r for r in results if r.get("status") == "ok"]
    has_fit = any(_fit(r).get("recomendacao") for r in ok)

    # Dashboard removido do Excel — ver dashboard.py (Streamlit, porta 8081)
    _sheet_candidatos(wb, ok, has_fit)
    if has_fit:
        _sheet_pontos_fortes(wb, ok)
        _sheet_lacunas_tab(wb, ok)
    _sheet_mapa(wb, ok, has_fit)
    if has_fit:
        _sheet_inteligencia(wb, ok, tema)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# Aba 1 — Dashboard (redesign completo)
# ══════════════════════════════════════════════════════════════════════════════
def _sheet_dashboard(wb, ok, tema, recrutador, empresa, has_fit):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False

    # Layout de colunas: A=margin | B-H=área esq | I=div | J-P=área dir | Q=margin
    col_widths = {1: 2, 2: 2}
    for c in range(3, 10):   col_widths[c] = 15   # C-I: área esquerda
    col_widths[10] = 2                              # J: divisor
    for c in range(11, 18):  col_widths[c] = 15   # K-Q: área direita
    col_widths[18] = 2
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── Cabeçalho ──
    ws.merge_cells("C2:Q2")
    c = ws["C2"]
    c.value = "CoreExtract — Dashboard Executivo de Triagem"
    c.font = _font(bold=True, size=20, color=_WHITE);  c.fill = _fill(_NAVY)
    c.alignment = _center();  ws.row_dimensions[2].height = 44

    ws.merge_cells("C3:Q3")
    c = ws["C3"]
    c.value = f"Vaga: {tema}  ·  {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    c.font = _font(size=11, color="A0B4CC", italic=True);  c.fill = _fill("0A1628")
    c.alignment = _center();  ws.row_dimensions[3].height = 22

    if recrutador or empresa:
        ws.merge_cells("C4:Q4")
        c = ws["C4"]
        parts = []
        if recrutador: parts.append(f"Recrutador: {recrutador}")
        if empresa:    parts.append(f"Empresa: {empresa}")
        c.value = "  ·  ".join(parts)
        c.font = _font(size=10, color=_GRAY);  c.fill = _fill(_SLATE)
        c.alignment = _center();  ws.row_dimensions[4].height = 18

    # ── KPI Cards (linha 6–7) ──
    total     = len(ok)
    rec_c     = _rec_count(ok)
    avg_score = round(sum(_score(r) for r in ok) / total) if total else 0
    pct_av    = f"{round(rec_c['Avançar']         / total * 100)}%" if total else "0%"
    pct_rej   = f"{round(rec_c['Não recomendado'] / total * 100)}%" if total else "0%"
    avg_lac   = round(sum(len(_fit(r).get("lacunas", [])) for r in ok) / total, 1) if total else 0

    if has_fit:
        cards = [
            ("Score Médio",         str(avg_score),  _PURPLE,   _WHITE,  "C", "D"),
            ("Total Candidatos",    str(total),       _NAVY,     _WHITE,  "E", "F"),
            ("% Avançar",           pct_av,           "007A4C",  _WHITE,  "G", "H"),
            ("% Rejeitados",        pct_rej,          "8B1A1A",  _WHITE,  "I", "J"),
            ("Lacunas / Candidato", str(avg_lac),     "5A3E8A",  _WHITE,  "K", "L"),
        ]
    else:
        nivel_counts = Counter(
            (_resume(r).get("nivel_senioridade") or "Indefinido") for r in ok
        )
        senior = nivel_counts.get("Sênior", 0) + nivel_counts.get("Especialista", 0)
        cards = [
            ("Score Médio",          str(avg_score),                       _PURPLE, _WHITE,  "C", "D"),
            ("Total Candidatos",     str(total),                            _NAVY,   _WHITE,  "E", "F"),
            ("Sênior / Especialista",str(senior),                          "007A4C", _WHITE,  "G", "H"),
            ("Pleno",                str(nivel_counts.get("Pleno", 0)),    "1A3A6B", _WHITE,  "I", "J"),
            ("Júnior",               str(nivel_counts.get("Júnior", 0)),  _GRAY,    _WHITE,  "K", "L"),
        ]

    for label, value, bg, fg, c1, c2 in cards:
        _kpi_card(ws, 6, c1, c2, label, value, bg, fg)
    ws.row_dimensions[6].height = 20
    ws.row_dimensions[7].height = 48
    ws.row_dimensions[8].height = 8

    # ── Nota dados ──
    ws.merge_cells("C9:Q9")
    c = ws["C9"]
    c.value = "⬇  Tabelas de dados para os gráficos — não edite as linhas abaixo"
    c.font = _font(size=8, color="C0CCE0", italic=True)
    ws.row_dimensions[9].height = 14

    # ─────────────────────────────────────────────────────────────────────────
    # TABELAS DE DADOS (col 3-9 para esquerda, col 11-17 para direita)
    # Gráficos ancorados em C10 e K10, depois C34 e K34
    # ─────────────────────────────────────────────────────────────────────────
    if has_fit:

        # ── Tabela A esquerda: Cross-tab Senioridade × Recomendação ──
        xtab = {n: {rec: 0 for rec in _REC_ORDER} for n in _NIVEL_ORDER}
        for r in ok:
            nivel = _resume(r).get("nivel_senioridade") or "Indefinido"
            if nivel not in xtab: nivel = "Indefinido"
            rec = _fit(r).get("recomendacao", "")
            if rec in _REC_ORDER:
                xtab[nivel][rec] += 1
        active_lvls = [n for n in _NIVEL_ORDER if sum(xtab[n].values()) > 0]
        n_lvls = max(len(active_lvls), 1)

        tA = 10
        ws.cell(row=tA, column=3, value="Nível").font              = _font(bold=True, size=9, color=_GRAY)
        ws.cell(row=tA, column=4, value="Avançar").font            = _font(bold=True, size=9, color="007A4C")
        ws.cell(row=tA, column=5, value="Em análise").font         = _font(bold=True, size=9, color="7A4F00")
        ws.cell(row=tA, column=6, value="Não recomendado").font    = _font(bold=True, size=9, color="8B1A1A")
        for i, nivel in enumerate(active_lvls):
            ws.cell(row=tA+1+i, column=3, value=nivel).font          = _font(size=9)
            ws.cell(row=tA+1+i, column=4, value=xtab[nivel]["Avançar"]).font         = _font(size=9)
            ws.cell(row=tA+1+i, column=5, value=xtab[nivel]["Em análise"]).font      = _font(size=9)
            ws.cell(row=tA+1+i, column=6, value=xtab[nivel]["Não recomendado"]).font = _font(size=9)

        # Gráfico 1 — Barra 100% Empilhada (Senioridade × Recomendação)
        _h1 = max(10, n_lvls * 2.5)
        chart1 = BarChart()
        chart1.type = "bar";  chart1.grouping = "percentStacked"
        chart1.title = "Recomendação por Senioridade"
        chart1.x_axis.title = "% dos Candidatos"
        chart1.style = 2;  chart1.width = 16;  chart1.height = _h1
        dr1 = Reference(ws, min_col=4, max_col=6, min_row=tA, max_row=tA + n_lvls)
        ct1 = Reference(ws, min_col=3, max_col=3, min_row=tA+1, max_row=tA + n_lvls)
        chart1.add_data(dr1, titles_from_data=True)
        chart1.set_categories(ct1)
        _set_bar_color(chart1.series[0], _GREEN,  _GREEN_D)   # Avançar
        _set_bar_color(chart1.series[1], _AMBER,  _AMBER_D)   # Em análise
        _set_bar_color(chart1.series[2], _RED,    _RED_D)     # Não recomendado
        ws.add_chart(chart1, "C10")

        # ── Tabela A direita: Score Ranking ──
        sorted_score = sorted(
            [(_cand_name(r, i), _score(r), _fit(r).get("recomendacao", ""))
             for i, r in enumerate(ok)],
            key=lambda x: -x[1],
        )
        n_cands = max(len(sorted_score), 1)

        for i, (name, sc, _rec) in enumerate(sorted_score):
            ws.cell(row=tA+1+i, column=11, value=name).font = _font(size=9)
            ws.cell(row=tA+1+i, column=12, value=sc).font   = _font(size=9)
        ws.cell(row=tA, column=11, value="Candidato").font  = _font(bold=True, size=9, color=_GRAY)
        ws.cell(row=tA, column=12, value="Score").font      = _font(bold=True, size=9, color=_GRAY)

        # Gráfico 2 — Ranking por Score, barras coloridas por recomendação
        _h2 = min(16, max(10, n_cands * 1.6))
        chart2 = BarChart()
        chart2.type = "bar";  chart2.grouping = "clustered"
        chart2.title = "Ranking por Score"
        chart2.x_axis.title = "Score (0–100)"
        chart2.style = 2;  chart2.width = 22;  chart2.height = _h2
        chart2.varyColors = False
        dr2 = Reference(ws, min_col=12, max_col=12, min_row=tA, max_row=tA + n_cands)
        ct2 = Reference(ws, min_col=11, max_col=11, min_row=tA+1, max_row=tA + n_cands)
        chart2.add_data(dr2, titles_from_data=True)
        chart2.set_categories(ct2)
        # Cada barra colorida pela recomendação do candidato
        _color_bar_by_rec(chart2.series[0], [(s, rec) for _, s, rec in sorted_score])
        ws.add_chart(chart2, "K10")

        # ── Tabela B esquerda: Pontos Fortes — frequência ──
        fortes_freq = _agg(ok, "pontos_fortes", top=10)
        nf = max(len(fortes_freq), 1)
        tB = tA + n_lvls + 3

        ws.cell(row=tB, column=3, value="Ponto Forte").font = _font(bold=True, size=9, color=_GRAY)
        ws.cell(row=tB, column=4, value="Freq.").font       = _font(bold=True, size=9, color=_GRAY)
        for i, (item, cnt) in enumerate(fortes_freq):
            ws.cell(row=tB+1+i, column=3, value=item[:52]).font = _font(size=9)
            ws.cell(row=tB+1+i, column=4, value=cnt).font       = _font(size=9)

        # Gráfico 3 — Pontos Fortes (verde)
        _h3 = min(16, max(10, nf * 1.4))
        row_c34 = 34   # âncora fixa para segunda linha de gráficos
        chart3 = BarChart()
        chart3.type = "bar";  chart3.grouping = "clustered"
        chart3.title = "Pontos Fortes — Frequência no Pool"
        chart3.x_axis.title = "Nº de Candidatos"
        chart3.style = 2;  chart3.width = 16;  chart3.height = _h3
        chart3.varyColors = False
        dr3 = Reference(ws, min_col=4, max_col=4, min_row=tB, max_row=tB + nf)
        ct3 = Reference(ws, min_col=3, max_col=3, min_row=tB+1, max_row=tB + nf)
        chart3.add_data(dr3, titles_from_data=True)
        chart3.set_categories(ct3)
        _set_bar_color(chart3.series[0], _GREEN, _GREEN_D)
        ws.add_chart(chart3, f"C{row_c34}")

        # ── Tabela B direita: Lacunas — frequência ──
        lacunas_freq = _agg(ok, "lacunas", top=10)
        nl = max(len(lacunas_freq), 1)
        tC = tB + nf + 3

        ws.cell(row=tC, column=11, value="Lacuna").font    = _font(bold=True, size=9, color=_GRAY)
        ws.cell(row=tC, column=12, value="Freq.").font     = _font(bold=True, size=9, color=_GRAY)
        for i, (item, cnt) in enumerate(lacunas_freq):
            ws.cell(row=tC+1+i, column=11, value=item[:52]).font = _font(size=9)
            ws.cell(row=tC+1+i, column=12, value=cnt).font       = _font(size=9)

        # Gráfico 4 — Lacunas (vermelho)
        _h4 = min(16, max(10, nl * 1.4))
        chart4 = BarChart()
        chart4.type = "bar";  chart4.grouping = "clustered"
        chart4.title = "Lacunas — Frequência no Pool"
        chart4.x_axis.title = "Nº de Candidatos"
        chart4.style = 2;  chart4.width = 22;  chart4.height = _h4
        chart4.varyColors = False
        dr4 = Reference(ws, min_col=12, max_col=12, min_row=tC, max_row=tC + nl)
        ct4 = Reference(ws, min_col=11, max_col=11, min_row=tC+1, max_row=tC + nl)
        chart4.add_data(dr4, titles_from_data=True)
        chart4.set_categories(ct4)
        _set_bar_color(chart4.series[0], _RED, _RED_D)
        ws.add_chart(chart4, f"K{row_c34}")

        # ── Bloco 5: Matriz Candidato × Competências ──
        top_skills = [item for item, _ in _agg(ok, "pontos_fortes", top=10)]
        row_mat = row_c34 + int(max(_h3, _h4) / 0.5) + 4

        ws.merge_cells(f"C{row_mat}:Q{row_mat}")
        c = ws[f"C{row_mat}"]
        c.value = "▶  MATRIZ CANDIDATO × COMPETÊNCIAS   ( ✓ = ponto forte identificado pela IA )"
        c.font = _font(bold=True, size=12, color=_WHITE);  c.fill = _fill("1A3A6B")
        c.alignment = _left(wrap=False);  ws.row_dimensions[row_mat].height = 28
        row_mat += 1

        # Cabeçalho da matriz
        cell = ws.cell(row=row_mat, column=3, value="Candidato")
        cell.fill = _fill(_NAVY);  cell.font = _font(bold=True, size=9, color=_WHITE)
        cell.alignment = _center();  cell.border = _border()
        for ci, skill in enumerate(top_skills, 4):
            cell = ws.cell(row=row_mat, column=ci, value=skill[:20])
            cell.fill = _fill("007A4C");  cell.font = _font(bold=True, size=8, color=_WHITE)
            cell.alignment = _center();  cell.border = _border()
        ws.row_dimensions[row_mat].height = 40
        row_mat += 1

        # Linhas: candidatos ordenados por recomendação → score desc
        sorted_mat = sorted(ok, key=lambda r: (
            _REC_ORDER.index(_fit(r).get("recomendacao", "Não recomendado")
                             if _fit(r).get("recomendacao") in _REC_ORDER
                             else "Não recomendado"),
            -_score(r),
        ))
        for i, r in enumerate(sorted_mat):
            nome = _cand_name(r, i)
            rec  = _fit(r).get("recomendacao", "")
            cand_set = {f.strip() for f in _fit(r).get("pontos_fortes", [])}
            row_bg = {"Avançar":"E8FBF1","Em análise":"FFF9E8","Não recomendado":"FFF0F0"}.get(rec,"F5F7FF")

            cell = ws.cell(row=row_mat, column=3, value=nome)
            cell.fill = _fill(row_bg);  cell.font = _font(bold=True, size=9)
            cell.alignment = _left(wrap=False);  cell.border = _border()

            for ci, skill in enumerate(top_skills, 4):
                present = skill in cand_set
                val  = "✓" if present else ""
                cell = ws.cell(row=row_mat, column=ci, value=val)
                cell.fill      = _fill("D6FAE8" if present else row_bg)
                cell.font      = _font(bold=True, size=11, color="007A4C") if present else _font(size=10, color="D0D8F0")
                cell.alignment = _center();  cell.border = _border()
            ws.row_dimensions[row_mat].height = 20
            row_mat += 1

    else:
        # ── Sem fit: pie de níveis + habilidades ──
        nivel_counts = Counter(
            (_resume(r).get("nivel_senioridade") or "Indefinido") for r in ok
        )
        tA = 10
        ws.cell(row=tA, column=3, value="Nível").font  = _font(bold=True, size=9, color=_GRAY)
        ws.cell(row=tA, column=4, value="Qtd").font    = _font(bold=True, size=9, color=_GRAY)
        for i, (nivel, cnt) in enumerate(nivel_counts.items()):
            ws.cell(row=tA+1+i, column=3, value=nivel).font = _font(size=9)
            ws.cell(row=tA+1+i, column=4, value=cnt).font   = _font(size=9)
        nn = len(nivel_counts)

        chart1 = PieChart()
        chart1.title = "Distribuição por Nível"
        chart1.style = 2;  chart1.width = 16;  chart1.height = 10
        dr = Reference(ws, min_col=4, max_col=4, min_row=tA, max_row=tA + nn)
        ct = Reference(ws, min_col=3, max_col=3, min_row=tA+1, max_row=tA + nn)
        chart1.add_data(dr, titles_from_data=True)
        chart1.set_categories(ct)
        ws.add_chart(chart1, "C10")

        tB = tA + nn + 3
        skills_freq = _all_skills(ok, top=10)
        ns = max(len(skills_freq), 1)
        ws.cell(row=tB, column=11, value="Habilidade").font  = _font(bold=True, size=9, color=_GRAY)
        ws.cell(row=tB, column=12, value="Candidatos").font  = _font(bold=True, size=9, color=_GRAY)
        for i, (skill, cnt) in enumerate(skills_freq):
            ws.cell(row=tB+1+i, column=11, value=skill).font = _font(size=9)
            ws.cell(row=tB+1+i, column=12, value=cnt).font   = _font(size=9)

        chart2 = BarChart()
        chart2.type = "bar";  chart2.grouping = "clustered"
        chart2.title = "Habilidades mais Frequentes no Pool"
        chart2.x_axis.title = "Candidatos"
        chart2.style = 2;  chart2.width = 22;  chart2.height = min(14, max(10, ns * 1.4))
        chart2.varyColors = False
        dr2 = Reference(ws, min_col=12, max_col=12, min_row=tB, max_row=tB + ns)
        ct2 = Reference(ws, min_col=11, max_col=11, min_row=tB+1, max_row=tB + ns)
        chart2.add_data(dr2, titles_from_data=True)
        chart2.set_categories(ct2)
        _set_bar_color(chart2.series[0], _BLUE, "0050CC")
        ws.add_chart(chart2, "K10")


# ══════════════════════════════════════════════════════════════════════════════
# Aba 2 — Candidatos
# ══════════════════════════════════════════════════════════════════════════════
def _sheet_candidatos(wb, ok, has_fit):
    ws = wb.create_sheet("Candidatos")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    if has_fit:
        cols = [
            ("Rank", 6), ("Nome", 26), ("Score", 8), ("Recomendação", 18),
            ("Pontos Fortes", 48), ("Lacunas", 48),
            ("Nível", 12), ("Área", 20), ("Exp. (anos)", 11),
            ("Habilidades Técnicas", 40), ("Idiomas", 22),
            ("E-mail", 30), ("Telefone", 16), ("LinkedIn", 34),
            ("Confiança IA", 12), ("Arquivo", 30),
        ]
    else:
        cols = [
            ("Rank", 6), ("Nome", 26), ("Score", 8),
            ("Nível", 12), ("Área", 20), ("Exp. (anos)", 11),
            ("Habilidades Técnicas", 40), ("Idiomas", 22),
            ("E-mail", 30), ("Telefone", 16), ("LinkedIn", 34),
            ("Confiança IA", 12), ("Obs. IA", 40), ("Arquivo", 30),
        ]

    for ci, (h, w) in enumerate(cols, 1):
        _header_cell(ws, 1, ci, h, w)
    ws.row_dimensions[1].height = 28

    for rank, r in enumerate(ok, 1):
        res  = _resume(r);  fit  = _fit(r)
        hb   = _habs(r);    ct   = _contato(r)
        row  = rank + 1
        rec  = fit.get("recomendacao", "") if has_fit else ""
        row_fill = _fill(_LIGHTBG) if rank % 2 == 0 else _fill(_WHITE)

        if has_fit:
            fortes  = " · ".join(fit.get("pontos_fortes", []))
            lacunas = " · ".join(fit.get("lacunas", []))
            values = [
                rank, res.get("nome", ""), _score(r), rec, fortes, lacunas,
                res.get("nivel_senioridade", ""), res.get("area_atuacao", ""),
                res.get("anos_experiencia_estimados", ""),
                ", ".join(hb.get("tecnicas", [])),
                ", ".join(hb.get("idiomas", [])),
                ct.get("email", ""), ct.get("telefone", ""), ct.get("linkedin", ""),
                res.get("confianca_extracao", ""), r.get("filename", ""),
            ]
        else:
            values = [
                rank, res.get("nome", ""), _score(r),
                res.get("nivel_senioridade", ""), res.get("area_atuacao", ""),
                res.get("anos_experiencia_estimados", ""),
                ", ".join(hb.get("tecnicas", [])),
                ", ".join(hb.get("idiomas", [])),
                ct.get("email", ""), ct.get("telefone", ""), ct.get("linkedin", ""),
                res.get("confianca_extracao", ""), res.get("observacoes_ia", ""),
                r.get("filename", ""),
            ]

        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.border = _border();  cell.font = _font(bold=(ci == 2), size=10)
            cell.alignment = _center() if ci in {1, 3, 7, 9} else _left()
            cell.fill = row_fill

            if has_fit and ci == 4 and val in _REC_STYLE:
                bg, fg = _REC_STYLE[val]
                cell.fill = _fill(bg);  cell.font = _font(bold=True, size=10, color=fg)
            nivel_col = 7 if has_fit else 4
            if ci == nivel_col and val in _NIVEL_STYLE:
                bg, fg = _NIVEL_STYLE[val]
                cell.fill = _fill(bg);  cell.font = _font(bold=True, size=10, color=fg)
        ws.row_dimensions[row].height = 30 if has_fit else 20

    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(ok)+1}"


# ══════════════════════════════════════════════════════════════════════════════
# Aba 3 — Pontos_Fortes (tabela normalizada — 1 linha por skill)
# ══════════════════════════════════════════════════════════════════════════════
def _sheet_pontos_fortes(wb, ok):
    ws = wb.create_sheet("Pontos_Fortes")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    cols = [("#", 6), ("ID Candidato", 14), ("Candidato", 28), ("Ponto Forte", 60)]
    for ci, (h, w) in enumerate(cols, 1):
        _header_cell(ws, 1, ci, h, w)
    ws.row_dimensions[1].height = 28

    row = 2
    for cand_id, r in enumerate(ok, 1):
        nome = _resume(r).get("nome", "") or r.get("filename", "")
        for skill in _fit(r).get("pontos_fortes", []):
            skill = skill.strip()
            if not skill:
                continue
            values = [row - 1, cand_id, nome, skill]
            row_fill = _fill("F0FFF8") if (row - 2) % 2 == 0 else _fill(_WHITE)
            for ci, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=ci, value=val)
                cell.border = _border();  cell.font = _font(size=10)
                cell.fill = row_fill
                cell.alignment = _center() if ci <= 2 else _left()
            ws.row_dimensions[row].height = 18
            row += 1

    if row > 2:
        ws.auto_filter.ref = f"A1:D{row-1}"


# ══════════════════════════════════════════════════════════════════════════════
# Aba 4 — Lacunas (tabela normalizada — 1 linha por lacuna)
# ══════════════════════════════════════════════════════════════════════════════
def _sheet_lacunas_tab(wb, ok):
    ws = wb.create_sheet("Lacunas")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    cols = [("#", 6), ("ID Candidato", 14), ("Candidato", 28), ("Lacuna", 60)]
    for ci, (h, w) in enumerate(cols, 1):
        _header_cell(ws, 1, ci, h, w)
    ws.row_dimensions[1].height = 28

    row = 2
    for cand_id, r in enumerate(ok, 1):
        nome = _resume(r).get("nome", "") or r.get("filename", "")
        for lacuna in _fit(r).get("lacunas", []):
            lacuna = lacuna.strip()
            if not lacuna:
                continue
            values = [row - 1, cand_id, nome, lacuna]
            row_fill = _fill("FFF5F5") if (row - 2) % 2 == 0 else _fill(_WHITE)
            for ci, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=ci, value=val)
                cell.border = _border();  cell.font = _font(size=10)
                cell.fill = row_fill
                cell.alignment = _center() if ci <= 2 else _left()
            ws.row_dimensions[row].height = 18
            row += 1

    if row > 2:
        ws.auto_filter.ref = f"A1:D{row-1}"


# ══════════════════════════════════════════════════════════════════════════════
# Aba 5 — Mapa de Candidatos
# ══════════════════════════════════════════════════════════════════════════════
def _sheet_mapa(wb, ok, has_fit):
    ws = wb.create_sheet("Mapa de Candidatos")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    cols = [
        ("Rank", 5), ("Nome", 26), ("Score", 8), ("Nível", 12),
        ("Área de Atuação", 22), ("Exp. (anos)", 11),
        ("Recomendação", 18) if has_fit else ("Status", 14),
        ("Pontos Fortes", 52) if has_fit else ("Habilidades Principais", 44),
        ("Lacunas", 52) if has_fit else ("Observações IA", 44),
        ("Habilidades Técnicas", 44),
        ("Idiomas", 22), ("Observações IA", 38), ("Arquivo", 28),
    ]

    for ci, (h, w) in enumerate(cols, 1):
        _header_cell(ws, 1, ci, h, w)
    ws.row_dimensions[1].height = 28

    for rank, r in enumerate(ok, 1):
        res = _resume(r);  fit = _fit(r);  hb = _habs(r)
        row = rank + 1
        rec = fit.get("recomendacao", "") if has_fit else ""

        if rec == "Avançar":            base_fill = _fill("E8FBF1")
        elif rec == "Em análise":       base_fill = _fill("FFF9E8")
        elif rec == "Não recomendado":  base_fill = _fill("FFF0F0")
        else:
            base_fill = _fill(_LIGHTBG) if rank % 2 == 0 else _fill(_WHITE)

        if has_fit:
            fortes  = " · ".join(fit.get("pontos_fortes", []))
            lacunas = " · ".join(fit.get("lacunas", []))
            values = [
                rank, res.get("nome",""), _score(r), res.get("nivel_senioridade",""),
                res.get("area_atuacao",""), res.get("anos_experiencia_estimados",""),
                rec, fortes, lacunas,
                ", ".join(hb.get("tecnicas",[])),
                ", ".join(hb.get("idiomas",[])),
                res.get("observacoes_ia",""), r.get("filename",""),
            ]
        else:
            values = [
                rank, res.get("nome",""), _score(r), res.get("nivel_senioridade",""),
                res.get("area_atuacao",""), res.get("anos_experiencia_estimados",""),
                "", ", ".join(hb.get("tecnicas",[])), res.get("observacoes_ia",""),
                ", ".join(hb.get("ferramentas",[])),
                ", ".join(hb.get("idiomas",[])),
                res.get("confianca_extracao",""), r.get("filename",""),
            ]

        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.border = _border();  cell.font = _font(bold=(ci==2), size=10)
            cell.fill   = base_fill
            cell.alignment = _center() if ci in {1,3,4,6} else _left()

            if has_fit and ci == 7 and val in _REC_STYLE:
                bg, fg = _REC_STYLE[val]
                cell.fill = _fill(bg);  cell.font = _font(bold=True, size=11, color=fg)
                cell.alignment = _center()
            if ci == 4 and val in _NIVEL_STYLE:
                bg, fg = _NIVEL_STYLE[val]
                cell.fill = _fill(bg);  cell.font = _font(bold=True, size=10, color=fg)

        ws.row_dimensions[row].height = 36 if has_fit else 24

    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(ok)+1}"


# ══════════════════════════════════════════════════════════════════════════════
# Aba 6 — Inteligência da Vaga  (só quando há fit)
# ══════════════════════════════════════════════════════════════════════════════
def _sheet_inteligencia(wb, ok, tema):
    ws = wb.create_sheet("Inteligência da Vaga")
    ws.sheet_view.showGridLines = False

    for i, w in enumerate([2, 28, 28, 12, 12, 12, 12, 2], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("B2:G2")
    c = ws["B2"]
    c.value = f"Inteligência da Vaga — {tema}"
    c.font = _font(bold=True, size=18, color=_WHITE);  c.fill = _fill(_NAVY)
    c.alignment = _center();  ws.row_dimensions[2].height = 36

    ws.merge_cells("B3:G3")
    c = ws["B3"]
    c.value = "Padrões identificados pela IA a partir da análise de fit de todos os candidatos"
    c.font = _font(size=10, color="A0B4CC", italic=True);  c.fill = _fill("0A1628")
    c.alignment = _center();  ws.row_dimensions[3].height = 20

    row = 5

    # ── Ranking ──
    ws.merge_cells(f"B{row}:G{row}")
    c = ws[f"B{row}"]
    c.value = "▶  RANKING DE CANDIDATOS"
    c.font = _font(bold=True, size=12, color=_WHITE);  c.fill = _fill("1A3A6B")
    c.alignment = _left(wrap=False);  ws.row_dimensions[row].height = 24
    row += 1

    rank_headers = ["Nome", "Score", "Recomendação", "Ponto Forte Principal", "Lacuna Principal", "Nível"]
    rh_widths    = [28, 8, 18, 40, 40, 12]
    for ci, (h, w) in enumerate(zip(rank_headers, rh_widths), 2):
        _header_cell(ws, row, ci, h, w)
    ws.row_dimensions[row].height = 22
    row += 1

    sorted_ok = sorted(ok, key=lambda r: (
        _REC_ORDER.index(_fit(r).get("recomendacao","Não recomendado")
                         if _fit(r).get("recomendacao") in _REC_ORDER
                         else "Não recomendado"),
        -_score(r),
    ))
    for r in sorted_ok:
        res = _resume(r);  fit = _fit(r)
        rec     = fit.get("recomendacao", "")
        fortes  = fit.get("pontos_fortes", [])
        lacunas = fit.get("lacunas", [])
        values  = [
            res.get("nome",""), _score(r), rec,
            fortes[0]  if fortes  else "—",
            lacunas[0] if lacunas else "—",
            res.get("nivel_senioridade",""),
        ]
        row_fill = (
            _fill("E8FBF1") if rec == "Avançar" else
            _fill("FFF9E8") if rec == "Em análise" else
            _fill("FFF0F0") if rec == "Não recomendado" else _fill(_WHITE)
        )
        for ci, val in enumerate(values, 2):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.border = _border();  cell.font = _font(size=10)
            cell.fill   = row_fill
            cell.alignment = _center() if ci in {3, 7} else _left()
            if ci == 4 and val in _REC_STYLE:
                bg, fg = _REC_STYLE[val]
                cell.fill = _fill(bg);  cell.font = _font(bold=True, size=10, color=fg)
                cell.alignment = _center()
        ws.row_dimensions[row].height = 24
        row += 1

    row += 1

    # ── Pontos Fortes Recorrentes ──
    ws.merge_cells(f"B{row}:G{row}")
    c = ws[f"B{row}"]
    c.value = "▶  PONTOS FORTES RECORRENTES NO POOL"
    c.font = _font(bold=True, size=12, color=_WHITE);  c.fill = _fill("007A4C")
    c.alignment = _left(wrap=False);  ws.row_dimensions[row].height = 24
    row += 1

    fortes_freq = _agg(ok, "pontos_fortes", top=10)
    if fortes_freq:
        ws.cell(row=row, column=2, value="Ponto Forte Identificado pela IA").font = _font(bold=True, size=10, color=_GRAY)
        ws.cell(row=row, column=6, value="Candidatos com este ponto").font        = _font(bold=True, size=10, color=_GRAY)
        ws.row_dimensions[row].height = 20;  row += 1
        for item, cnt in fortes_freq:
            pct = f"{round(cnt/len(ok)*100)}% do pool" if ok else ""
            ws.cell(row=row, column=2, value=f"✓  {item}").font          = _font(size=10, color="007A4C")
            ws.cell(row=row, column=6, value=f"{cnt} ({pct})").font      = _font(bold=True, size=10, color="007A4C")
            for ci in [2, 6]:
                ws.cell(row=row, column=ci).fill      = _fill("F0FFF8")
                ws.cell(row=row, column=ci).border    = _border()
                ws.cell(row=row, column=ci).alignment = _left(wrap=False)
            ws.row_dimensions[row].height = 20;  row += 1
    else:
        ws.cell(row=row, column=2, value="Sem pontos fortes identificados.").font = _font(size=10, italic=True, color=_GRAY)
        row += 1

    row += 1

    # ── Lacunas Recorrentes ──
    ws.merge_cells(f"B{row}:G{row}")
    c = ws[f"B{row}"]
    c.value = "▶  LACUNAS RECORRENTES NO POOL  (habilidades que o mercado tem dificuldade de suprir)"
    c.font = _font(bold=True, size=12, color=_WHITE);  c.fill = _fill("8B1A1A")
    c.alignment = _left(wrap=False);  ws.row_dimensions[row].height = 24
    row += 1

    lacunas_freq = _agg(ok, "lacunas", top=10)
    if lacunas_freq:
        ws.cell(row=row, column=2, value="Lacuna Identificada pela IA").font = _font(bold=True, size=10, color=_GRAY)
        ws.cell(row=row, column=6, value="Candidatos afetados").font         = _font(bold=True, size=10, color=_GRAY)
        ws.row_dimensions[row].height = 20;  row += 1
        for item, cnt in lacunas_freq:
            pct      = f"{round(cnt/len(ok)*100)}% do pool" if ok else ""
            severity = "⚠️  CRÍTICO" if cnt >= len(ok) * 0.6 else ("⚡  COMUM" if cnt >= 2 else "")
            label    = f"✗  {item}" + (f"   [{severity}]" if severity else "")
            ws.cell(row=row, column=2, value=label).font          = _font(size=10, color="CC2200")
            ws.cell(row=row, column=6, value=f"{cnt} ({pct})").font = _font(bold=True, size=10, color="CC2200")
            for ci in [2, 6]:
                ws.cell(row=row, column=ci).fill      = _fill("FFF5F5")
                ws.cell(row=row, column=ci).border    = _border()
                ws.cell(row=row, column=ci).alignment = _left(wrap=False)
            ws.row_dimensions[row].height = 20;  row += 1
    else:
        ws.cell(row=row, column=2, value="Sem lacunas identificadas.").font = _font(size=10, italic=True, color=_GRAY)
        row += 1

    row += 1

    # ── Habilidades Técnicas do Pool ──
    ws.merge_cells(f"B{row}:G{row}")
    c = ws[f"B{row}"]
    c.value = "▶  HABILIDADES TÉCNICAS DISPONÍVEIS NO POOL"
    c.font = _font(bold=True, size=12, color=_WHITE);  c.fill = _fill(_BLUE)
    c.alignment = _left(wrap=False);  ws.row_dimensions[row].height = 24
    row += 1

    skills = _all_skills(ok, top=16)
    if skills:
        col_a, col_b = 2, 5
        ws.cell(row=row, column=col_a,   value="Habilidade").font = _font(bold=True, size=10, color=_GRAY)
        ws.cell(row=row, column=col_a+1, value="# Cands").font   = _font(bold=True, size=10, color=_GRAY)
        ws.cell(row=row, column=col_b,   value="Habilidade").font = _font(bold=True, size=10, color=_GRAY)
        ws.cell(row=row, column=col_b+1, value="# Cands").font   = _font(bold=True, size=10, color=_GRAY)
        ws.row_dimensions[row].height = 20;  row += 1

        half = (len(skills) + 1) // 2
        left_skills  = skills[:half]
        right_skills = skills[half:]
        for i in range(half):
            if i < len(left_skills):
                s, c_val = left_skills[i]
                ws.cell(row=row+i, column=col_a,   value=s).font     = _font(size=10)
                ws.cell(row=row+i, column=col_a+1, value=c_val).font = _font(bold=True, size=10)
                ws.cell(row=row+i, column=col_a).fill     = _fill("F0F4FF")
                ws.cell(row=row+i, column=col_a+1).fill   = _fill("F0F4FF")
                ws.cell(row=row+i, column=col_a).border   = _border()
                ws.cell(row=row+i, column=col_a+1).border = _border()
                ws.cell(row=row+i, column=col_a+1).alignment = _center()
            if i < len(right_skills):
                s, c_val = right_skills[i]
                ws.cell(row=row+i, column=col_b,   value=s).font     = _font(size=10)
                ws.cell(row=row+i, column=col_b+1, value=c_val).font = _font(bold=True, size=10)
                ws.cell(row=row+i, column=col_b).fill     = _fill("F0F4FF")
                ws.cell(row=row+i, column=col_b+1).fill   = _fill("F0F4FF")
                ws.cell(row=row+i, column=col_b).border   = _border()
                ws.cell(row=row+i, column=col_b+1).border = _border()
                ws.cell(row=row+i, column=col_b+1).alignment = _center()
            ws.row_dimensions[row+i].height = 18
